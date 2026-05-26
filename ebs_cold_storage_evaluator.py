#!/usr/bin/env python3
"""
EBS Snapshot Cold Storage Evaluator
Evaluates cold storage eligibility at the VOLUME level using the EBS Direct API
(ListChangedBlocks / ListSnapshotBlocks) to measure actual unique block ratios.

Single threshold: 25% — archive saves money when unreferenced blocks >= 25% of full snapshot.

Math:
  Standard tier:  $0.05/GB-month   (incremental only)
  Archive tier:   $0.0125/GB-month (full snapshot)
  Break-even:     unreferenced_size >= full_snapshot_size * (0.0125 / 0.05) = 25%
"""

import boto3
import json
import csv
import sys
from datetime import datetime, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

BLOCK_SIZE_BYTES = 524288  # 512 KiB
ARCHIVE_MIN_DAYS = 90
THRESHOLD_PCT = 25  # Single unambiguous threshold


def get_ebs_pricing(session, region):
    """Get EBS snapshot pricing for Standard and Archive tiers."""
    ssm = session.client("ssm", region_name="us-east-1")
    try:
        param = ssm.get_parameter(
            Name=f"/aws/service/global-infrastructure/regions/{region}/longName"
        )
        location = param["Parameter"]["Value"]
    except Exception:
        location = region

    pricing_client = session.client("pricing", region_name="us-east-1")
    std_price = 0.05
    try:
        resp = pricing_client.get_products(
            ServiceCode="AmazonEC2",
            Filters=[
                {"Type": "TERM_MATCH", "Field": "productFamily", "Value": "Storage Snapshot"},
                {"Type": "TERM_MATCH", "Field": "storageMedia", "Value": "Amazon S3"},
                {"Type": "TERM_MATCH", "Field": "location", "Value": location},
            ],
            MaxResults=10,
        )
        for item in resp["PriceList"]:
            product = json.loads(item)
            for term in product.get("terms", {}).get("OnDemand", {}).values():
                for dim in term.get("priceDimensions", {}).values():
                    price = float(dim["pricePerUnit"]["USD"])
                    if price > 0:
                        std_price = price
                        break
    except Exception:
        pass
    return std_price, std_price * 0.25  # Archive is always 25% of standard


def get_snapshots(ec2_client):
    """Get all completed standard-tier snapshots owned by this account."""
    snapshots = []
    paginator = ec2_client.get_paginator("describe_snapshots")
    for page in paginator.paginate(
        Filters=[
            {"Name": "status", "Values": ["completed"]},
            {"Name": "storage-tier", "Values": ["standard"]},
        ],
        OwnerIds=["self"],
    ):
        snapshots.extend(page["Snapshots"])
    return snapshots


def get_backup_expiration(session):
    """Get expiration info from AWS Backup recovery points."""
    backup_client = session.client("backup")
    snap_expiration = {}
    try:
        vaults = backup_client.list_backup_vaults().get("BackupVaultList", [])
        for vault in vaults:
            paginator = backup_client.get_paginator("list_recovery_points_by_backup_vault")
            for page in paginator.paginate(BackupVaultName=vault["BackupVaultName"]):
                for rp in page.get("RecoveryPoints", []):
                    if rp.get("ResourceType") != "EBS":
                        continue
                    rp_arn = rp.get("RecoveryPointArn", "")
                    if "/snap-" not in rp_arn:
                        continue
                    snap_id = rp_arn.split("/")[-1]
                    delete_at = rp.get("CalculatedLifecycle", {}).get("DeleteAt")
                    if delete_at:
                        snap_expiration[snap_id] = {
                            "delete_at": delete_at,
                            "days_left": (delete_at - datetime.now(delete_at.tzinfo)).days,
                            "retention_days": rp.get("Lifecycle", {}).get("DeleteAfterDays"),
                        }
    except Exception as e:
        print(f"  Warning: Could not fully query AWS Backup: {e}")
    return snap_expiration


def check_volume_exists(ec2_client, volume_id):
    """Check if the source volume still exists."""
    if volume_id in ("vol-ffffffff", "unknown"):
        return False
    try:
        ec2_client.describe_volumes(VolumeIds=[volume_id])
        return True
    except Exception:
        return False


def count_changed_blocks(ebs_client, first_snap_id, second_snap_id):
    """Count changed blocks between two snapshots using ListChangedBlocks."""
    count = 0
    token = None
    while True:
        kwargs = {
            "FirstSnapshotId": first_snap_id,
            "SecondSnapshotId": second_snap_id,
        }
        if token:
            kwargs["NextToken"] = token
        resp = ebs_client.list_changed_blocks(**kwargs)
        count += len(resp.get("ChangedBlocks", []))
        token = resp.get("NextToken")
        if not token:
            break
    return count


def count_snapshot_blocks(ebs_client, snapshot_id):
    """Count total blocks in a snapshot using ListSnapshotBlocks."""
    count = 0
    token = None
    while True:
        kwargs = {"SnapshotId": snapshot_id}
        if token:
            kwargs["NextToken"] = token
        resp = ebs_client.list_snapshot_blocks(**kwargs)
        count += len(resp.get("Blocks", []))
        token = resp.get("NextToken")
        if not token:
            break
    return count


def get_unreferenced_blocks(ebs_client, snap_id, prev_snap_id, next_snap_id):
    """
    Find unreferenced blocks for a snapshot using ListChangedBlocks.
    Unreferenced = blocks that changed from predecessor AND changed in successor.
    These blocks are unique to this snapshot and won't be re-attributed.
    """
    # Blocks changed between predecessor and this snapshot
    if prev_snap_id:
        changed_from_prev = set()
        token = None
        while True:
            kwargs = {"FirstSnapshotId": prev_snap_id, "SecondSnapshotId": snap_id}
            if token:
                kwargs["NextToken"] = token
            resp = ebs_client.list_changed_blocks(**kwargs)
            for block in resp.get("ChangedBlocks", []):
                changed_from_prev.add(block["BlockIndex"])
            token = resp.get("NextToken")
            if not token:
                break
    else:
        # First snapshot in lineage — all its blocks are "changed from nothing"
        changed_from_prev = None

    # Blocks changed between this snapshot and successor
    if next_snap_id:
        changed_to_next = set()
        token = None
        while True:
            kwargs = {"FirstSnapshotId": snap_id, "SecondSnapshotId": next_snap_id}
            if token:
                kwargs["NextToken"] = token
            resp = ebs_client.list_changed_blocks(**kwargs)
            for block in resp.get("ChangedBlocks", []):
                changed_to_next.add(block["BlockIndex"])
            token = resp.get("NextToken")
            if not token:
                break
    else:
        # Last snapshot in lineage — none of its blocks are referenced by a successor
        changed_to_next = None

    # Unreferenced = in both sets (unique to this snapshot, not shared with neighbors)
    if changed_from_prev is None and changed_to_next is None:
        # Only snapshot — all blocks are unreferenced
        return None  # Signal: use full snapshot size
    elif changed_from_prev is None:
        # First in lineage — unreferenced = blocks overwritten by successor
        return len(changed_to_next)
    elif changed_to_next is None:
        # Last in lineage — unreferenced = blocks it changed from predecessor
        return len(changed_from_prev)
    else:
        # Middle — unreferenced = blocks in BOTH sets
        return len(changed_from_prev & changed_to_next)


def evaluate_volumes(region, profile=None, max_retention_days=None):
    """Main evaluation at VOLUME level using EBS Direct API."""
    session = boto3.Session(profile_name=profile, region_name=region)
    ec2 = session.client("ec2")
    ebs = session.client("ebs")

    print(f"\n{'='*70}")
    print(f"EBS Snapshot Cold Storage Evaluator (Volume-Level, 25% Threshold)")
    print(f"Region: {region} | Profile: {profile or 'default'}")
    print(f"{'='*70}")

    # Pricing
    print("\n[1/4] Fetching EBS pricing...")
    std_price, archive_price = get_ebs_pricing(session, region)
    print(f"  Standard: ${std_price}/GB-month | Archive: ${archive_price}/GB-month")
    print(f"  Break-even: unreferenced blocks >= 25% of full snapshot")

    # Snapshots
    print("\n[2/4] Discovering snapshots...")
    snapshots = get_snapshots(ec2)
    print(f"  Found {len(snapshots)} completed standard-tier snapshots")
    if not snapshots:
        print("  No snapshots found. Exiting.")
        return

    # Expiration
    print("\n[3/4] Checking expiration dates...")
    snap_expiration = get_backup_expiration(session)
    print(f"  Found expiration info for {len(snap_expiration)} snapshots")

    # Group by volume
    vol_snaps = {}
    for snap in snapshots:
        vol_id = snap.get("VolumeId", "unknown")
        vol_snaps.setdefault(vol_id, []).append(snap)
    for vol_id in vol_snaps:
        vol_snaps[vol_id].sort(key=lambda s: s["StartTime"])
    print(f"  {len(vol_snaps)} unique volumes")

    # Evaluate
    print("\n[4/4] Evaluating volumes (EBS Direct API)...")
    volume_results = []

    for vol_id, snaps in vol_snaps.items():
        print(f"\n  Volume: {vol_id} ({len(snaps)} snapshots)")
        vol_size_gb = snaps[0]["VolumeSize"]
        oldest_snap = snaps[0]
        newest_snap = snaps[-1]
        span_days = (newest_snap["StartTime"] - oldest_snap["StartTime"]).days

        volume_exists = check_volume_exists(ec2, vol_id)

        # Check expiration
        expired_count = 0
        active_expiry_count = 0
        min_expiry_days = None
        for s in snaps:
            sid = s["SnapshotId"]
            if sid in snap_expiration:
                days_left = snap_expiration[sid]["days_left"]
                if days_left < 0:
                    expired_count += 1
                else:
                    active_expiry_count += 1
                    if min_expiry_days is None or days_left < min_expiry_days:
                        min_expiry_days = days_left

        # Check max retention period (user-specified)
        exceeds_retention_count = 0
        if max_retention_days is not None:
            now = datetime.now(oldest_snap["StartTime"].tzinfo)
            for s in snaps:
                age_days = (now - s["StartTime"]).days
                if age_days > max_retention_days:
                    exceeds_retention_count += 1

        # Get full snapshot size (newest)
        newest_full_bytes = int(newest_snap.get("FullSnapshotSizeInBytes", 0))
        newest_full_gb = newest_full_bytes / (1024**3) if newest_full_bytes else vol_size_gb

        # === DECISION LOGIC ===
        recommendation = ""
        reason = ""
        unreferenced_pct = None

        # Case 0: Snapshots exceed max retention period (user-specified)
        if max_retention_days is not None and exceeds_retention_count == len(snaps):
            recommendation = "HOUSEKEEP - EXCEEDS RETENTION"
            reason = (f"All {len(snaps)} snapshots exceed max retention of {max_retention_days} days. "
                      f"Consider deleting to save ~${newest_full_gb * std_price:.2f}/month.")

        elif max_retention_days is not None and exceeds_retention_count > 0:
            recommendation = "HOUSEKEEP - EXCEEDS RETENTION"
            reason = (f"{exceeds_retention_count}/{len(snaps)} snapshots exceed max retention of "
                      f"{max_retention_days} days. Consider deleting over-retained snapshots.")

        # Case 1: All snapshots expired
        if expired_count == len(snaps):
            recommendation = "HOUSEKEEP - DELETE ALL"
            reason = (f"All {len(snaps)} snapshots have expired retention. "
                      f"Delete to save ~${newest_full_gb * std_price:.2f}/month.")

        # Case 2: Snapshots expire within 90 days
        elif min_expiry_days is not None and min_expiry_days < ARCHIVE_MIN_DAYS:
            recommendation = "NOT ELIGIBLE"
            reason = (f"Snapshots expire in {min_expiry_days} days "
                      f"(< {ARCHIVE_MIN_DAYS}-day archive minimum).")

        # Case 3: Mixed expired + active
        elif expired_count > 0:
            recommendation = "HOUSEKEEP FIRST"
            reason = (f"{expired_count}/{len(snaps)} snapshots expired. "
                      f"Delete expired first, then re-evaluate.")

        # Case 4: Volume decommissioned
        elif not volume_exists:
            recommendation = "COLD STORAGE CANDIDATE"
            reason = (f"Volume decommissioned. No new snapshots will be created. "
                      f"All {len(snaps)} snapshots can be archived — no re-attribution issue. "
                      f"Saves ~75% (${newest_full_gb * std_price:.2f} → "
                      f"${newest_full_gb * archive_price:.2f}/month).")

        # Case 5: Single snapshot
        elif len(snaps) == 1:
            recommendation = "COLD STORAGE CANDIDATE"
            reason = (f"Single standalone snapshot — no lineage sharing. "
                      f"Archive saves 75% (${newest_full_gb * std_price:.2f} → "
                      f"${newest_full_gb * archive_price:.2f}/month).")

        # Case 6: Multiple snapshots — use EBS Direct API
        else:
            try:
                # Evaluate the newest snapshot (most likely archive candidate)
                # For a complete evaluation, check each snapshot individually
                snap_idx = len(snaps) - 1  # newest
                snap_id = snaps[snap_idx]["SnapshotId"]
                prev_snap_id = snaps[snap_idx - 1]["SnapshotId"] if snap_idx > 0 else None
                next_snap_id = snaps[snap_idx + 1]["SnapshotId"] if snap_idx < len(snaps) - 1 else None

                # Get full block count for this snapshot
                full_blocks = count_snapshot_blocks(ebs, snap_id)

                # Get unreferenced block count
                unreferenced = get_unreferenced_blocks(ebs, snap_id, prev_snap_id, next_snap_id)

                if unreferenced is None:
                    # Only snapshot (shouldn't reach here, but safety)
                    unreferenced_pct = 100.0
                elif full_blocks > 0:
                    unreferenced_pct = (unreferenced / full_blocks) * 100
                else:
                    unreferenced_pct = 0.0

                if unreferenced_pct >= THRESHOLD_PCT:
                    recommendation = "COLD STORAGE CANDIDATE"
                    reason = (f"Unreferenced blocks = {unreferenced_pct:.1f}% of full snapshot "
                              f"(>= {THRESHOLD_PCT}% threshold). "
                              f"Archiving saves money. "
                              f"Unreferenced: {unreferenced} blocks, Full: {full_blocks} blocks.")
                else:
                    recommendation = "NOT RECOMMENDED"
                    reason = (f"Unreferenced blocks = {unreferenced_pct:.1f}% of full snapshot "
                              f"(< {THRESHOLD_PCT}% threshold). "
                              f"Archiving would likely INCREASE costs due to re-attribution. "
                              f"Unreferenced: {unreferenced} blocks, Full: {full_blocks} blocks.")

            except Exception as e:
                recommendation = "ERROR - MANUAL REVIEW"
                reason = f"Could not query EBS Direct API: {e}"

        volume_results.append({
            "region": region,
            "volume_id": vol_id,
            "volume_size_gb": vol_size_gb,
            "volume_exists": "Yes" if volume_exists else "No (Decommissioned)",
            "total_snapshots": len(snaps),
            "oldest_snapshot_id": oldest_snap["SnapshotId"],
            "oldest_snapshot_date": oldest_snap["StartTime"].isoformat(),
            "newest_snapshot_id": newest_snap["SnapshotId"],
            "newest_snapshot_date": newest_snap["StartTime"].isoformat(),
            "snapshot_span_days": span_days,
            "newest_full_snapshot_gb": round(newest_full_gb, 2),
            "unreferenced_pct": f"{unreferenced_pct:.1f}%" if unreferenced_pct is not None else "N/A",
            "expired_snapshots": expired_count,
            "exceeds_retention": exceeds_retention_count if max_retention_days else "N/A",
            "min_days_to_expiry": min_expiry_days if min_expiry_days is not None else "No expiry",
            "warm_cost_per_month_est": round(newest_full_gb * std_price, 2),
            "cold_cost_per_month_est": round(newest_full_gb * archive_price, 2),
            "recommendation": recommendation,
            "decision_reason": reason,
        })
        print(f"    → {recommendation}")

    # Output
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_filename = f"ebs_cold_storage_eval_{region}_{timestamp}.csv"

    if volume_results:
        with open(csv_filename, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=volume_results[0].keys())
            writer.writeheader()
            writer.writerows(volume_results)
        print(f"\n  CSV: {csv_filename}")

    if HAS_OPENPYXL and volume_results:
        xlsx_filename = f"ebs_cold_storage_eval_{region}_{timestamp}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Cold Storage Assessment"
        headers = list(volume_results[0].keys())
        ws.append(headers)

        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        orange_fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
        red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

        for row_data in volume_results:
            ws.append([row_data[h] for h in headers])
            rec = row_data["recommendation"]
            if "COLD STORAGE CANDIDATE" in rec:
                fill = green_fill
            elif "EXCEEDS RETENTION" in rec:
                fill = yellow_fill
            elif "HOUSEKEEP" in rec:
                fill = orange_fill
            elif rec in ("NOT RECOMMENDED", "NOT ELIGIBLE"):
                fill = red_fill
            else:
                continue
            for cell in ws[ws.max_row]:
                cell.fill = fill

        # --- Cost Savings Summary Sheet ---
        ss = wb.create_sheet("Cost Savings Summary")
        ss.append(["Category", "Volumes", "Snapshots", "Monthly Warm Cost", "Monthly Cold Cost", "Monthly Savings"])

        candidates = [r for r in volume_results if "COLD STORAGE CANDIDATE" in r["recommendation"]]
        housekeep = [r for r in volume_results if "HOUSEKEEP" in r["recommendation"]]
        retention_exceeded = [r for r in volume_results if "EXCEEDS RETENTION" in r["recommendation"]]

        cand_warm = sum(r["warm_cost_per_month_est"] for r in candidates)
        cand_cold = sum(r["cold_cost_per_month_est"] for r in candidates)
        cand_snaps = sum(r["total_snapshots"] for r in candidates)

        hk_warm = sum(r["warm_cost_per_month_est"] for r in housekeep)
        hk_snaps = sum(r["total_snapshots"] for r in housekeep)

        ret_warm = sum(r["warm_cost_per_month_est"] for r in retention_exceeded)
        ret_snaps = sum(r["exceeds_retention"] for r in retention_exceeded if r["exceeds_retention"] != "N/A")

        ss.append(["Archive to Cold Storage", len(candidates), cand_snaps,
                   f"${cand_warm:.2f}", f"${cand_cold:.2f}", f"${cand_warm - cand_cold:.2f}"])
        ss.append(["Housekeep (Delete Expired)", len(housekeep), hk_snaps,
                   f"${hk_warm:.2f}", "$0.00", f"${hk_warm:.2f}"])
        if max_retention_days:
            ss.append(["Exceeds Retention (Delete)", len(retention_exceeded), int(ret_snaps),
                       f"${ret_warm:.2f}", "$0.00", f"${ret_warm:.2f}"])

        total_savings = (cand_warm - cand_cold) + hk_warm + ret_warm
        ss.append([])
        ss.append(["TOTAL POTENTIAL SAVINGS", "", "", "", "", f"${total_savings:.2f}/month"])
        ss.append(["ANNUAL SAVINGS", "", "", "", "", f"${total_savings * 12:.2f}/year"])

        # Bold the totals
        from openpyxl.styles import Font
        bold = Font(bold=True)
        for cell in ss[ss.max_row]:
            cell.font = bold
        for cell in ss[ss.max_row - 1]:
            cell.font = bold

        wb.save(xlsx_filename)
        print(f"  Excel: {xlsx_filename}")

    # Summary
    print(f"\n{'='*70}")
    print(f"SUMMARY — {region}")
    print(f"{'='*70}")
    print(f"  Volumes evaluated: {len(volume_results)}")
    candidates = [r for r in volume_results if "COLD STORAGE CANDIDATE" in r["recommendation"]]
    housekeep = [r for r in volume_results if "HOUSEKEEP" in r["recommendation"]]
    not_rec = [r for r in volume_results if r["recommendation"] in ("NOT RECOMMENDED", "NOT ELIGIBLE")]
    print(f"  ✅ Cold storage candidates: {len(candidates)}")
    print(f"  🟠 Housekeep/delete first:  {len(housekeep)}")
    print(f"  ❌ Not recommended/eligible: {len(not_rec)}")
    if candidates:
        print(f"\n  COLD STORAGE CANDIDATES:")
        for c in candidates:
            print(f"    {c['volume_id']} | {c['total_snapshots']} snaps | "
                  f"unreferenced: {c['unreferenced_pct']} | {c['recommendation']}")
    print(f"{'='*70}\n")
    return volume_results


if __name__ == "__main__":
    profile = sys.argv[1] if len(sys.argv) > 1 else None
    regions = sys.argv[2].split(",") if len(sys.argv) > 2 else ["us-east-1"]
    max_retention_days = int(sys.argv[3]) if len(sys.argv) > 3 else None

    all_results = []
    for region in regions:
        region = region.strip()
        results = evaluate_volumes(region=region, profile=profile, max_retention_days=max_retention_days)
        if results:
            all_results.extend(results)

    # Combined output across all regions
    if all_results:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        region_label = "_".join(r.strip() for r in regions) if len(regions) <= 3 else f"{len(regions)}_regions"
        csv_filename = f"ebs_cold_storage_eval_{region_label}_{timestamp}.csv"
        with open(csv_filename, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=all_results[0].keys())
            writer.writeheader()
            writer.writerows(all_results)
        print(f"\n  Combined CSV: {csv_filename}")

        if HAS_OPENPYXL:
            xlsx_filename = f"ebs_cold_storage_eval_{region_label}_{timestamp}.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Cold Storage Assessment"
            headers = list(all_results[0].keys())
            ws.append(headers)

            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            orange_fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
            red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

            for row_data in all_results:
                ws.append([row_data[h] for h in headers])
                rec = row_data["recommendation"]
                if "COLD STORAGE CANDIDATE" in rec:
                    fill = green_fill
                elif "EXCEEDS RETENTION" in rec:
                    fill = yellow_fill
                elif "HOUSEKEEP" in rec:
                    fill = orange_fill
                elif rec in ("NOT RECOMMENDED", "NOT ELIGIBLE"):
                    fill = red_fill
                else:
                    continue
                for cell in ws[ws.max_row]:
                    cell.fill = fill

            # --- Cost Savings Summary Sheet ---
            ss = wb.create_sheet("Cost Savings Summary")
            ss.append(["Category", "Volumes", "Snapshots", "Monthly Warm Cost", "Monthly Cold Cost", "Monthly Savings"])

            candidates = [r for r in all_results if "COLD STORAGE CANDIDATE" in r["recommendation"]]
            housekeep = [r for r in all_results if "HOUSEKEEP" in r["recommendation"]]
            retention_exceeded = [r for r in all_results if "EXCEEDS RETENTION" in r["recommendation"]]

            cand_warm = sum(r["warm_cost_per_month_est"] for r in candidates)
            cand_cold = sum(r["cold_cost_per_month_est"] for r in candidates)
            cand_snaps = sum(r["total_snapshots"] for r in candidates)

            hk_warm = sum(r["warm_cost_per_month_est"] for r in housekeep)
            hk_snaps = sum(r["total_snapshots"] for r in housekeep)

            ret_warm = sum(r["warm_cost_per_month_est"] for r in retention_exceeded)
            ret_snaps = sum(r["exceeds_retention"] for r in retention_exceeded if r["exceeds_retention"] != "N/A")

            ss.append(["Archive to Cold Storage", len(candidates), cand_snaps,
                       f"${cand_warm:.2f}", f"${cand_cold:.2f}", f"${cand_warm - cand_cold:.2f}"])
            ss.append(["Housekeep (Delete Expired)", len(housekeep), hk_snaps,
                       f"${hk_warm:.2f}", "$0.00", f"${hk_warm:.2f}"])
            if max_retention_days:
                ss.append(["Exceeds Retention (Delete)", len(retention_exceeded), int(ret_snaps),
                           f"${ret_warm:.2f}", "$0.00", f"${ret_warm:.2f}"])

            total_savings = (cand_warm - cand_cold) + hk_warm + ret_warm
            ss.append([])
            ss.append(["TOTAL POTENTIAL SAVINGS", "", "", "", "", f"${total_savings:.2f}/month"])
            ss.append(["ANNUAL SAVINGS", "", "", "", "", f"${total_savings * 12:.2f}/year"])

            from openpyxl.styles import Font
            bold = Font(bold=True)
            for cell in ss[ss.max_row]:
                cell.font = bold
            for cell in ss[ss.max_row - 1]:
                cell.font = bold

            wb.save(xlsx_filename)
            print(f"  Combined Excel: {xlsx_filename}")
